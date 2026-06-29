"""
OEE ETL Pipeline
================
Reads daily IML OEE Excel files and produces a single JSON file
that the OEE dashboard (oee_dashboard_v2.html) reads directly.

REQUIREMENTS
------------
Install once, in Command Prompt:
    py -m pip install openpyxl pandas

USAGE
-----
Place all your daily Excel files in one folder, then run:

    py oee_etl.py

By default it looks for Excel files in the same folder as this script
and writes oee_data.json next to it.

To specify a different folder:
    py oee_etl.py --folder "D:\WORKING\OEE\\data\\OEE Files"

To write the JSON somewhere else:
    py oee_etl.py --output "D:\WORKING\OEE\\output\\oee_data.json"

FILE NAMING
-----------
Each Excel file must contain the date somewhere in its name as DD_MM_YYYY.
Examples that all work:
    21_03_2026.xlsx
    OEE_21_03_2026.xlsx
    1774355103920_21_03_2026.xlsx

FORMULA METHODOLOGY
--------------------
All metrics follow the Sheet1 methodology from your Excel workbook exactly:

  Shift run %       = machines that ran / total machine-slots  (from Shift tracker)
  Availability loss = (sum shift_hrs - sum run_hrs) / sum shift_hrs
  Cavity loss       = (sum total_cav - sum running_cav) / sum total_cav
  Quality loss      = sum rejected_pcs / (sum rejected + sum good)
  OEE (running)     = sum actual_good / sum target             (Sheet1 R6)
  Speed loss        = residual: OEE_running / (avail x cav x quality)
  Overall OEE       = OEE_running x shift_run_pct              (Sheet1 D10)

  Downtime %        = each category hrs / total shift hrs      (Sheet2 methodology)
  Not-run reasons   = count per reason / total machine-slots
"""

import os
import re
import sys
import json
import argparse
import logging
from datetime import datetime
from pathlib import Path

# ── Dependency check (friendly error before import fails) ────────────────────
missing = []
try:
    import pandas as pd
except ImportError:
    missing.append("pandas")
try:
    from openpyxl import load_workbook
except ImportError:
    missing.append("openpyxl")

if missing:
    print("\nERROR: Missing required libraries. Please run this in Command Prompt:")
    print(f"\n    py -m pip install {' '.join(missing)}\n")
    sys.exit(1)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("oee_etl")

# ── Column positions in OEE tracker (1-indexed, for openpyxl) ────────────────
# Two layouts supported — detected automatically from row 10 header:
#
#   OLD layout (col I = Product type):   no Operator name column
#   NEW layout (col I = Operator name):  inserted at col 9, all others +1
#
# Three template layouts exist:
#   OLDEST — no Operator column, OEE% at col AP (42), 8 DT categories
#   OLD    — Operator at col I (9),  OEE% at col AQ (43), 8 DT categories
#   NEW    — Operator at col I (9),  OEE% at col AT (46), 11 DT categories
#
# Detection order:
#   1. Col AT (46) non-empty at row 10 → NEW
#   2. Col I  (9)  header contains "Operator" → OLD
#   3. Otherwise → OLDEST

OEE_COL_OLDEST = {
    "date":             2,   # B
    "shift":            3,   # C
    "duration_hrs":     6,   # F
    "machine_no":       7,   # G
    "to_fill":          8,   # H
    # No operator_name column in this template
    "product_type":     9,   # I  (where Operator is in OLD/NEW)
    "product_name":     10,  # J
    "mould_no":         11,  # K
    # L (12): No. of SKUs manufactured — skip
    "weight_gm":        13,  # M
    "rated_cycle_sec":  14,  # N
    "total_cavities":   15,  # O
    "target_pcs":       16,  # P
    "actual_good_pcs":  17,  # Q
    # R (18): SKU level production breakdown — skip
    "rej_label_kg":     19,  # S
    "rej_plain_kg":     20,  # T
    "rej_trial_kg":     21,  # U
    "total_rej_pcs":    22,  # V
    "std_weight_kg":     23,  # W  Total std material weight (kg, incl. rejected)
    # W (23): Total std material weight — skip
    # X (24): % target achievement (pre-computed) — skip
    # Y (25): Quality loss % (pre-computed) — skip
    "run_hrs":          26,  # Z
    "dt_label":         27,  # AA
    "dt_colour":        28,  # AB
    "dt_mould":         29,  # AC
    "dt_label_unavail": 30,  # AD
    "dt_proc_fail":     31,  # AE
    "dt_mach_bkdn":     32,  # AF  (combined machine+robot in this template)
    "dt_mould_bkdn":    33,  # AG
    "dt_other":         34,  # AH
    # AI (35): Comments — skip
    "avail_loss_pct":   36,  # AJ
    "running_cavities": 37,  # AK
    "perf_cav_loss":    38,  # AL
    "actual_cycle_sec": 39,  # AM
    "perf_spd_loss":    40,  # AN
    "overall_perf_loss":41,  # AO
    "oee_pct":          42,  # AP
}

OEE_COL_OLD = {
    "date":             2,   # B
    "shift":            3,   # C
    "duration_hrs":     6,   # F
    "machine_no":       7,   # G
    "to_fill":          8,   # H
    "operator_name":    9,   # I
    "product_type":     10,  # J
    "product_name":     11,  # K
    "mould_no":         12,  # L
    "weight_gm":        14,  # N
    "rated_cycle_sec":  15,  # O
    "total_cavities":   16,  # P
    "target_pcs":       17,  # Q
    "actual_good_pcs":  18,  # R
    # col 19 (S) = SKU-level production breakdown — not needed in ETL
    "rej_label_kg":     20,  # T
    "rej_plain_kg":     21,  # U
    "rej_trial_kg":     22,  # V
    "total_rej_pcs":    23,  # W
    "std_weight_kg":     24,  # X  Total std material weight (kg, incl. rejected)
    # cols 24-26 (X-Z) = material weight, % target, quality % — not needed
    "run_hrs":          27,  # AA
    "dt_label":         28,  # AB
    "dt_colour":        29,  # AC
    "dt_mould":         30,  # AD
    "dt_label_unavail": 31,  # AE
    "dt_proc_fail":     32,  # AF
    "dt_mach_bkdn":     33,  # AG  (combined machine+robot in OLD template)
    "dt_mould_bkdn":    34,  # AH
    "dt_other":         35,  # AI
    # col 36 (AJ): Comments for other reasons — skip
    "avail_loss_pct":   37,  # AK
    "running_cavities": 38,  # AL
    "perf_cav_loss":    39,  # AM
    "actual_cycle_sec": 40,  # AN
    "perf_spd_loss":    41,  # AO
    "overall_perf_loss":42,  # AP
    "oee_pct":          43,  # AQ
}

OEE_COL_NEW = {
    "date":             2,   # B
    "shift":            3,   # C
    "duration_hrs":     6,   # F
    "machine_no":       7,   # G
    "to_fill":          8,   # H
    "operator_name":    9,   # I
    "product_type":     10,  # J
    "product_name":     11,  # K
    "mould_no":         12,  # L
    "weight_gm":        14,  # N
    "rated_cycle_sec":  15,  # O
    "total_cavities":   16,  # P
    "target_pcs":       17,  # Q
    "actual_good_pcs":  18,  # R
    # col 19 (S) = SKU-level production breakdown — not needed in ETL
    "rej_label_kg":     20,  # T
    "rej_plain_kg":     21,  # U
    "rej_trial_kg":     22,  # V
    "total_rej_pcs":    23,  # W
    "std_weight_kg":     24,  # X  Total std material weight (kg, incl. rejected)
    # cols 24-26 (X-Z) = material weight, % target, quality % — not needed
    "run_hrs":          27,  # AA
    "dt_label":         28,  # AB
    "dt_colour":        29,  # AC
    "dt_mould":         30,  # AD
    "dt_label_unavail": 31,  # AE
    "dt_proc_fail":     32,  # AF
    "dt_mach_bkdn":     33,  # AG  ← machine breakdown (split from combined OLD)
    "dt_mould_bkdn":    34,  # AH
    "dt_robot_bkdn":    35,  # AI  ← robot breakdown (split from combined OLD)
    "dt_manpower":      36,  # AJ  ← manpower unavailability
    "dt_power_cut":     37,  # AK  ← power cut
    "dt_other":         38,  # AL
    # col 39 (AM): Comments for other reasons — skip
    "avail_loss_pct":   40,  # AN
    "running_cavities": 41,  # AO
    "perf_cav_loss":    42,  # AP
    "actual_cycle_sec": 43,  # AQ
    "perf_spd_loss":    44,  # AR
    "overall_perf_loss":45,  # AS
    "oee_pct":          46,  # AT
}

# v3 template: 3 extra DT columns added (startup/setup, RM-PM unavail, label peel-off)
# All post-DT columns shift right by 3; OEE% moves from AT(46) → AW(49)
OEE_COL_NEWEST = {
    "date":             2,   # B
    "shift":            3,   # C
    "duration_hrs":     6,   # F
    "machine_no":       7,   # G
    "to_fill":          8,   # H
    "operator_name":    9,   # I
    "product_type":     10,  # J
    "product_name":     11,  # K
    "mould_no":         12,  # L
    "weight_gm":        14,  # N
    "rated_cycle_sec":  15,  # O
    "total_cavities":   16,  # P
    "target_pcs":       17,  # Q
    "actual_good_pcs":  18,  # R
    # col 19 (S) = SKU-level breakdown — skip
    "rej_label_kg":     20,  # T
    "rej_plain_kg":     21,  # U
    "rej_trial_kg":     22,  # V
    "total_rej_pcs":    23,  # W
    "std_weight_kg":     24,  # X  Total std material weight (kg, incl. rejected)
    # cols 24-26 (X-Z) = material weight, % target, quality % — skip
    "run_hrs":          27,  # AA
    "dt_label":         28,  # AB
    "dt_colour":        29,  # AC
    "dt_mould":         30,  # AD
    "dt_startup":       31,  # AE  ← NEW: machine startup / setup
    "dt_label_unavail": 32,  # AF
    "dt_rm_pm_unavail": 33,  # AG  ← NEW: other RM/PM unavailability
    "dt_peel_off":      34,  # AH  ← NEW: label peel-off
    "dt_proc_fail":     35,  # AI
    "dt_mach_bkdn":     36,  # AJ
    "dt_mould_bkdn":    37,  # AK
    "dt_robot_bkdn":    38,  # AL
    "dt_manpower":      39,  # AM
    "dt_power_cut":     40,  # AN
    "dt_other":         41,  # AO
    # col 42 (AP): Comments — skip
    "avail_loss_pct":   43,  # AQ
    "running_cavities": 44,  # AR
    "perf_cav_loss":    45,  # AS
    "actual_cycle_sec": 46,  # AT
    "perf_spd_loss":    47,  # AU
    "overall_perf_loss":48,  # AV
    "oee_pct":          49,  # AW
}

def detect_oee_col(ws) -> dict:
    """
    Detect which of the four OEE tracker layouts this file uses.

    OLDEST:  no Operator column; col I = Product type; OEE% at AP (42); 8 DT cols
    OLD:     Operator at col I;  OEE% at AQ (43);  8 DT cols
    NEW:     Operator at col I;  OEE% at AT (46); 11 DT cols
    NEWEST:  Operator at col I;  OEE% at AW (49); 14 DT cols (adds startup, RM/PM unavail, peel-off)

    Detection order (most specific first):
      1. Col AW (49) header contains "OEE"  → NEWEST
      2. Col AT (46) non-empty              → NEW
      3. Col I  (9)  header contains "operator" → OLD
      4. Otherwise                          → OLDEST
    """
    aw_hdr = str(ws.cell(row=10, column=49).value or "").strip().lower()
    if "oee" in aw_hdr:
        log.info("  Detected NEWEST column layout (OEE% at col AW/49, 14 DT categories)")
        return OEE_COL_NEWEST
    if ws.cell(row=10, column=46).value not in (None, ""):
        log.info("  Detected NEW column layout (OEE% at col AT/46, 11 DT categories)")
        return OEE_COL_NEW
    col_i_hdr = str(ws.cell(row=10, column=9).value or "").strip().lower()
    if "operator" in col_i_hdr:
        log.info("  Detected OLD column layout (OEE% at col AQ/43, Operator at col I)")
        return OEE_COL_OLD
    log.info("  Detected OLDEST column layout (OEE% at col AP/42, no Operator column)")
    return OEE_COL_OLDEST

# OEE_COL is set per-file inside parse_file() — do not use this global directly
OEE_COL = OEE_COL_OLD  # fallback default (unused after parse_file sets it)

SHIFT_COL = {
    "date":      2,  # B
    "shift":     3,  # C
    "machine":   4,  # D
    "ran":       5,  # E
    "reason":    6,  # F
    "hours_run": 7,  # G  ← new column (Y=12, N=0, P=partial hours entered)
}

DT_KEYS   = ["dt_label","dt_colour","dt_mould","dt_startup",
             "dt_label_unavail","dt_rm_pm_unavail","dt_peel_off",
             "dt_proc_fail","dt_mach_bkdn","dt_mould_bkdn","dt_robot_bkdn",
             "dt_manpower","dt_power_cut","dt_other"]
DT_LABELS = ["Label change","Colour change","Mould change","Startup/setup",
             "Label unavail.","RM/PM unavail.","Label peel-off",
             "Process failure","Mach. breakdown","Mould breakdown","Robot breakdown",
             "Manpower unavail.","Power cut","Other"]


# ── Date extraction ───────────────────────────────────────────────────────────

def extract_date(path):
    """
    Find a date in the filename and return as datetime.
    Accepts any of these formats:
        DD_MM_YYYY   e.g. 21_03_2026
        DD.MM.YYYY   e.g. 21.03.2026
        DD-MM-YYYY   e.g. 21-03-2026
        DD MM YYYY   e.g. 21 03 2026
        D.MM.YYYY    e.g. 12.3.2026  (single-digit day or month)
        D.M.YYYY     e.g. 1.3.2026
    """
    stem = Path(path).stem
    # Try DD[sep]MM[sep]YYYY where sep is . _ - or space, digits can be 1 or 2
    m = re.search(r"(\d{1,2})[._\- ](\d{1,2})[._\- ](\d{4})", stem)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    log.warning(f"Could not parse date from: {Path(path).name}")
    return None


# ── Excel parsing ─────────────────────────────────────────────────────────────

def parse_file(path):
    """
    Parse one daily Excel file.
    Returns dict with keys: date_str, all_slots, not_run, records
    or None if the file cannot be parsed.
    """
    file_date = extract_date(path)
    if file_date is None:
        return None

    date_str = file_date.strftime("%Y-%m-%d")
    log.info(f"Reading: {Path(path).name}  ->  {date_str}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        log.error(f"  Cannot open file: {e}")
        return None

    if "Shift tracker" not in wb.sheetnames or "OEE tracker" not in wb.sheetnames:
        log.error(f"  Missing required sheets. Found: {wb.sheetnames}")
        return None

    # ── Shift tracker ──────────────────────────────────────────────────────
    ws_shift = wb["Shift tracker"]
    not_run  = []
    all_slots = []

    for row in range(15, 300):
        machine = ws_shift.cell(row=row, column=SHIFT_COL["machine"]).value
        shift   = ws_shift.cell(row=row, column=SHIFT_COL["shift"]).value
        ran     = ws_shift.cell(row=row, column=SHIFT_COL["ran"]).value

        if machine is None or shift is None:
            if row > 20:
                break
            continue

        try:
            machine_no = int(machine)
        except (ValueError, TypeError):
            continue

        shift_str  = str(shift).strip()
        ran_str    = str(ran).strip().upper() if ran is not None else "Y"
        reason_raw = ws_shift.cell(row=row, column=SHIFT_COL["reason"]).value
        reason_str = str(reason_raw).strip() if reason_raw else None

        # Y/N are definitive — ignore col G to prevent manual entry errors.
        # P relies on col G (hours actually run entered by supervisor).
        if ran_str == "Y":
            hours_run = 12.0
        elif ran_str == "N":
            hours_run = 0.0
        else:  # P — partial: read from col G
            hrs_raw = ws_shift.cell(row=row, column=SHIFT_COL["hours_run"]).value
            try:
                hours_run = float(hrs_raw) if hrs_raw is not None else 0.0
            except (ValueError, TypeError):
                hours_run = 0.0

        slot = {
            "date":       date_str,
            "shift":      shift_str,
            "machine_no": machine_no,
            "hours_run":  hours_run,
            "ran_flag":   ran_str,   # Y / N / P
        }
        all_slots.append(slot)

        if ran_str in ("N", "P"):
            not_run.append({
                "date":       date_str,
                "shift":      shift_str,
                "machine_no": machine_no,
                "hours_run":  hours_run,   # 0 for N; partial hrs for P
                "reason":     reason_str or "Missing reason code",
            })

    log.info(f"  Shift tracker: {len(all_slots)} slots, {len(not_run)} not run")

    # ── OEE tracker ────────────────────────────────────────────────────────
    ws_oee  = wb["OEE tracker"]
    OEE_COL = detect_oee_col(ws_oee)  # auto-detect old vs new layout
    records = []

    # ── Find last row with data so we don't break early ───────────────────
    # Scan once to find where data ends (last row with any value in key cols)
    last_row = 11
    for _r in range(11, 2000):
        has_data = any(
            ws_oee.cell(row=_r, column=c).value is not None
            for c in [2, 7, 8, 16, 17]  # date, machine, to_fill, target, actual
        )
        if has_data:
            last_row = _r

    for row in range(11, last_row + 1):
        # ── Row inclusion logic — matches Excel's SUMIF(O>0) exactly ──────
        # Include a row if rated_cycle_sec > 0 (column O), same filter as
        # the Excel aggregate formulas in the manual calculation sheet.
        # This covers rows with actual production = 0 (machine ran but no
        # good output) as long as a valid cycle time exists.
        def get(col_name, row=row):
            return ws_oee.cell(row=row, column=OEE_COL[col_name]).value

        def num(col_name, default=0, row=row):
            # Gracefully return default for columns not in this layout (old files)
            if col_name not in OEE_COL:
                return default
            v = ws_oee.cell(row=row, column=OEE_COL[col_name]).value
            try:
                return float(v) if v is not None else default
            except (ValueError, TypeError):
                return default

        cycle_sec = num("rated_cycle_sec")

        # Skip rows with no valid cycle time (VLOOKUP failed or row is empty)
        if cycle_sec <= 0:
            continue

        machine_no = num("machine_no")
        if machine_no == 0:
            continue

        records.append({
            "date":             date_str,
            "shift":            str(get("shift") or "").strip(),
            "machine_no":       int(machine_no),
            "operator_name":    str(get("operator_name") or "").strip() if "operator_name" in OEE_COL else "",
            "product_type":     str(get("product_type") or "").strip(),
            "product_name":     str(get("product_name") or "").strip(),
            "mould_no":         str(get("mould_no") or "").strip(),
            "duration_hrs":     num("duration_hrs"),
            "run_hrs":          num("run_hrs"),
            "total_cavities":   int(num("total_cavities")),
            "running_cavities": min(int(num("running_cavities")), int(num("total_cavities"))),  # cap at total — data entry errors can give rc>tc
            "target_pcs":       num("target_pcs"),
            "actual_good_pcs":  num("actual_good_pcs"),
            "total_rej_pcs":    num("total_rej_pcs"),
            "std_weight_kg":    num("std_weight_kg"),
            "rej_label_kg":     num("rej_label_kg"),
            "rej_plain_kg":     num("rej_plain_kg"),
            "rej_trial_kg":     num("rej_trial_kg"),
            "dt_label":         num("dt_label"),
            "dt_colour":        num("dt_colour"),
            "dt_mould":         num("dt_mould"),
            "dt_startup":       num("dt_startup"),        # v3: machine startup/setup
            "dt_label_unavail": num("dt_label_unavail"),
            "dt_rm_pm_unavail": num("dt_rm_pm_unavail"),  # v3: other RM/PM unavailability
            "dt_peel_off":      num("dt_peel_off"),       # v3: label peel-off
            "dt_proc_fail":     num("dt_proc_fail"),
            "dt_mach_bkdn":     num("dt_mach_bkdn"),
            "dt_mould_bkdn":    num("dt_mould_bkdn"),
            "dt_robot_bkdn":    num("dt_robot_bkdn"),
            "dt_manpower":      num("dt_manpower"),
            "dt_power_cut":     num("dt_power_cut"),
            "dt_other":         num("dt_other"),
        })

    log.info(f"  OEE tracker:   {len(records)} active machine-shift rows")

    # not_run is already populated from shift tracker N entries above.
    # The shift tracker is the source of truth — do not override with
    # "no OEE record" logic, which incorrectly flags Y slots on days
    # where OEE data was missing.
    log.info(f"  Not-run slots: {len(not_run)} (Shift tracker N entries)")

    return {
        "date_str":  date_str,
        "all_slots": all_slots,
        "not_run":   not_run,
        "records":   records,
    }


# ── Formula engine — matches Sheet1 exactly ───────────────────────────────────

def compute_metrics(records, not_run, all_slots):
    """
    Duration-weighted average approach: each metric is a per-row rate averaged
    by duration_hrs weight so every machine-hour counts equally regardless of
    product type. Speed loss is derived as residual.
    """
    def s(fn):
        return sum(fn(r) for r in records)

    total_dur_w = s(lambda r: r["duration_hrs"]) or 1

    def wt(r, val):
        return val * (r["duration_hrs"] or 0)

    avail_loss = s(lambda r: wt(r,
        (r["duration_hrs"] - r["run_hrs"]) / r["duration_hrs"]
        if r["duration_hrs"] else 0)) / total_dur_w

    cav_loss = s(lambda r: wt(r,
        (r["total_cavities"] - r["running_cavities"]) / r["total_cavities"]
        if r["total_cavities"] else 0)) / total_dur_w

    quality_loss = s(lambda r: wt(r,
        r["total_rej_pcs"] / (r["total_rej_pcs"] + r["actual_good_pcs"])
        if (r["total_rej_pcs"] + r["actual_good_pcs"]) else 0)) / total_dur_w

    oee_running = s(lambda r: wt(r,
        r["actual_good_pcs"] / r["target_pcs"]
        if r["target_pcs"] else 0)) / total_dur_w

    # Raw sums for output display fields
    total_shift_hrs = s(lambda r: r["duration_hrs"])
    total_run_hrs   = s(lambda r: r["run_hrs"])
    total_target    = s(lambda r: r["target_pcs"])
    total_actual    = s(lambda r: r["actual_good_pcs"])
    total_rej       = s(lambda r: r["total_rej_pcs"])

    # D4: shift run %
    # Total slots = all machine-shifts in Shift tracker (ground truth)
    # Ran = unique (date, shift, machine) with target_pcs > 0 in OEE tracker
    # Not run = total_slots - ran  (delta)
    total_slots   = len(all_slots)
    ran_keys_set  = {(r["date"], r["shift"], r["machine_no"]) for r in records}
    ran_cnt       = len(ran_keys_set)
    not_ran_cnt   = total_slots - ran_cnt

    # Pro-rated shift run % using hours_run from Shift tracker.
    # Y → 12 hrs, N → 0 hrs, P → partial hours (user entered).
    # For old files without hours_run, the fallback in parse_file sets
    # Y=12 / N=0, so the formula reduces to the old binary ran_cnt/total.
    SHIFT_HRS     = 12.0   # standard shift duration
    total_ran_hrs = sum(s.get("hours_run", SHIFT_HRS) for s in all_slots)
    shift_run_pct = total_ran_hrs / (total_slots * SHIFT_HRS) if total_slots else 1

    # Speed loss: residual — guarantees avail × cav × speed × qual = oee_running
    avail_rate   = 1 - avail_loss
    cav_rate     = 1 - cav_loss
    quality_rate = 1 - quality_loss
    denom        = avail_rate * cav_rate * quality_rate
    speed_rate   = oee_running / denom if denom > 0 else 1.0
    speed_loss   = max(0.0, 1 - speed_rate)

    # D10: overall OEE = OEE_running x shift_run_pct
    overall_oee = oee_running * shift_run_pct

    # Downtime breakdown as % of total shift hrs (Sheet2 methodology)
    dt_breakdown_pct = {}
    dt_breakdown_hrs = {}
    for k, label in zip(DT_KEYS, DT_LABELS):
        hrs = s(lambda r, k=k: r[k])
        dt_breakdown_pct[label] = round(hrs / total_shift_hrs, 6) if total_shift_hrs else 0
        dt_breakdown_hrs[label] = round(hrs, 2)

    # Not-run reasons as % of total slots
    reason_counts = {}
    for nr in not_run:
        key = nr["reason"] or "Missing reason code"
        reason_counts[key] = reason_counts.get(key, 0) + 1
    reason_pcts = {
        k: round(v / total_slots, 6) if total_slots else 0
        for k, v in reason_counts.items()
    }

    return {
        "total_shift_hrs":  round(total_shift_hrs, 1),
        "total_run_hrs":    round(total_run_hrs, 1),
        "total_target":     int(total_target),
        "total_actual":     int(total_actual),
        "total_rej":        round(total_rej, 0),
        "total_slots":      total_slots,
        "ran_slots":        ran_cnt,
        "not_ran_slots":    not_ran_cnt,
        "shift_run_pct":    round(shift_run_pct, 6),
        "avail_loss":       round(avail_loss, 6),
        "cav_loss":         round(cav_loss, 6),
        "speed_loss":       round(speed_loss, 6),
        "quality_loss":     round(quality_loss, 6),
        "oee_running":      round(oee_running, 6),
        "overall_oee":      round(overall_oee, 6),
        "dt_breakdown_pct": dt_breakdown_pct,
        "dt_breakdown_hrs": dt_breakdown_hrs,
        "reason_pcts":      reason_pcts,
        "reason_counts":    reason_counts,
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def _parse_int(v):
    """
    Robustly convert a cell value to int.
    Used for machine_no, cavities, and other fields that are always pure integers.
    Returns None if the value is empty or unparseable.
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))       # handles '12', '12.0', 12, 12.0
    except (ValueError, TypeError):
        pass
    m = re.match(r'^(\d+)', s)    # grabs leading digits from strings like '12abc'
    return int(m.group(1)) if m else None


def _mould_str(v) -> "str | None":
    """
    Convert a mould number cell value to a clean string, PRESERVING any suffix.
    Examples:
      19172       → '19172'
      19172.0     → '19172'     (Excel stores integers as floats)
      '19172(11)' → '19172(11)' (suffix kept — this is the full mould name)
      ' 19172 '   → '19172'     (whitespace stripped)
    Returns None if the value is empty.
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Excel numeric cells come through as floats (e.g. 19172.0) — strip trailing .0
    if re.match(r'^\d+\.0$', s):
        return s[:-2]
    return s


def export_master(excel_files: list, out_path: str) -> None:
    """
    Read the Master sheet and write oee_master.json.
    Priority:
      1. OEE_Master.xlsx in the same folder as the first excel_file (standalone master)
      2. Most recent daily Excel that contains a "Master" sheet (legacy fallback)
    Schema per combination:
      { product_name, weight_gm, mould_no, cavities, machine_no, cycle_time_sec, combination }
    """
    master_ws = None

    # ── 1. Look for standalone OEE_Master.xlsx first ────────────────────────
    if excel_files:
        standalone = Path(excel_files[0]).parent / "OEE_Master.xlsx"
        if standalone.exists():
            try:
                wb = load_workbook(str(standalone), data_only=True)
                if "Master" in wb.sheetnames:
                    master_ws = wb["Master"]
                    log.info(f"  Reading Master sheet from standalone {standalone.name}")
            except Exception as e:
                log.warning(f"  Could not read {standalone.name}: {e}")

    # ── 2. Fallback: most recent daily Excel with a Master sheet ─────────────
    if master_ws is None:
        for fp in reversed(excel_files):
            try:
                wb = load_workbook(str(fp), data_only=True)
                if "Master" in wb.sheetnames:
                    master_ws = wb["Master"]
                    log.info(f"  Reading Master sheet from {Path(fp).name} (fallback)")
                    break
            except Exception:
                continue

    if master_ws is None:
        log.warning("  No Master sheet found — oee_master.json not updated")
        return

    combinations = []
    seen = set()
    for row in range(3, master_ws.max_row + 1):
        product_name  = master_ws.cell(row=row, column=2).value
        weight_gm     = master_ws.cell(row=row, column=3).value
        mould_no      = master_ws.cell(row=row, column=4).value
        cavities      = master_ws.cell(row=row, column=5).value
        machine_no    = master_ws.cell(row=row, column=6).value
        cycle_time    = master_ws.cell(row=row, column=8).value

        if not product_name or not mould_no or not machine_no:
            continue

        pname     = str(product_name).strip()
        mould_s   = _mould_str(mould_no)
        machine_i = _parse_int(machine_no)

        if not mould_s or machine_i is None:
            log.warning(f"  Skipping master row {row}: could not parse mould_no={mould_no!r} or machine_no={machine_no!r}")
            continue

        key = (pname, mould_s, machine_i)
        if key in seen:
            continue
        seen.add(key)

        try:
            combinations.append({
                "product_name":    pname,
                "weight_gm":       float(weight_gm) if weight_gm else None,
                "mould_no":        mould_s,          # stored as string to preserve suffixes like '(11)'
                "cavities":        _parse_int(cavities),
                "machine_no":      machine_i,
                "cycle_time_sec":  float(cycle_time) if cycle_time else None,
                "combination":     f"{pname} | {mould_s} | {machine_i}",
            })
        except (ValueError, TypeError):
            continue

    # Derive product_type from name (LID → Lid, else Container)
    for c in combinations:
        c["product_type"] = "Lid" if "LID" in c["product_name"].upper() else "Container"

    # Unique sorted lists for dropdowns
    # Moulds are strings (may have suffixes like '19172(11)') — sort by numeric prefix first
    def _mould_sort_key(s):
        m = re.match(r'^(\d+)', str(s))
        return (int(m.group(1)) if m else 0, str(s))

    products = sorted(set(c["product_name"] for c in combinations))
    moulds   = sorted(set(c["mould_no"]     for c in combinations), key=_mould_sort_key)
    machines = sorted(set(c["machine_no"]   for c in combinations))

    master_data = {
        "_meta":        {"generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        "combinations": combinations,
        "products":     products,
        "moulds":       moulds,
        "machines":     machines,
    }

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(master_data, f, indent=2)
    log.info(f"  Master exported: {len(combinations)} combinations → {out_path}")


def read_json_submissions(folder: Path) -> list:
    """
    Read all JSON submission files from submissions/ subfolder.
    Returns a list of parse_file()-compatible result dicts.
    """
    sub_folder = folder / "submissions"
    if not sub_folder.exists():
        return []

    results = []
    for jf in sorted(sub_folder.glob("*.json")):
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
            # Each submission file contains a list of shift records for one date
            if not isinstance(data, dict) or "date" not in data:
                continue

            date_str  = data["date"]
            records   = data.get("records",   [])
            not_run   = data.get("not_run",   [])
            all_slots = data.get("all_slots", [])

            if not records and not all_slots:
                continue

            results.append({
                "date_str":  date_str,
                "records":   records,
                "not_run":   not_run,
                "all_slots": all_slots,
            })
            log.info(f"  Read JSON submission: {jf.name}  ({len(records)} OEE records, {len(all_slots)} slots)")
        except Exception as e:
            log.warning(f"  Could not read submission {jf.name}: {e}")

    return results


def main():
    parser = argparse.ArgumentParser(
        description="IML OEE ETL -- converts daily Excel files to dashboard JSON"
    )
    parser.add_argument(
        "--folder",
        default=".",
        help='Folder containing daily Excel files (default: same folder as this script)'
    )
    parser.add_argument(
        "--output",
        default="oee_data.json",
        help="Output JSON path (default: oee_data.json)"
    )
    parser.add_argument(
        "--master",
        default="oee_master.json",
        help="Master export JSON path (default: oee_master.json)"
    )
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists():
        log.error(f"Folder not found: {folder}")
        sys.exit(1)

    excel_files = sorted(
        list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")),
        key=lambda p: extract_date(str(p)) or datetime.min
    )
    log.info(f"Found {len(excel_files)} Excel file(s) in {folder.resolve()}")

    if not excel_files:
        log.error("No .xlsx files found. Check the --folder path.")
        sys.exit(1)

    # ── Export master data (always, from most recent Excel with Master sheet) ─
    log.info("Exporting master data...")
    export_master(excel_files, args.master)

    # ── Parse all files ────────────────────────────────────────────────────
    all_records  = []
    all_not_run  = []
    all_slots    = []
    daily_data   = []
    failed_files = []

    for fp in excel_files:
        result = parse_file(str(fp))
        if result is None:
            failed_files.append(fp.name)
            continue

        all_records.extend(result["records"])
        all_not_run.extend(result["not_run"])
        all_slots.extend(result["all_slots"])

        day_metrics = compute_metrics(
            result["records"],
            result["not_run"],
            result["all_slots"],
        )
        daily_data.append({
            "date":          result["date_str"],
            "overall_oee":   day_metrics["overall_oee"],
            "oee_running":   day_metrics["oee_running"],
            "shift_run_pct": day_metrics["shift_run_pct"],
            "avail_loss":    day_metrics["avail_loss"],
            "cav_loss":      day_metrics["cav_loss"],
            "speed_loss":    day_metrics["speed_loss"],
            "quality_loss":  day_metrics["quality_loss"],
            "total_actual":  day_metrics["total_actual"],
            "total_target":  day_metrics["total_target"],
        })

    # ── Read JSON submissions from submissions/ folder ─────────────────────
    log.info("Checking for JSON submissions...")
    json_results = read_json_submissions(folder)
    for jr in json_results:
        # Avoid double-counting: skip if same date already read from Excel
        existing_dates = {d["date"] for d in daily_data}
        if jr["date_str"] in existing_dates:
            log.info(f"  Skipping JSON for {jr['date_str']} — already read from Excel")
            continue
        all_records.extend(jr["records"])
        all_not_run.extend(jr["not_run"])
        all_slots.extend(jr["all_slots"])
        day_metrics = compute_metrics(jr["records"], jr["not_run"], jr["all_slots"])
        daily_data.append({
            "date":          jr["date_str"],
            "overall_oee":   day_metrics["overall_oee"],
            "oee_running":   day_metrics["oee_running"],
            "shift_run_pct": day_metrics["shift_run_pct"],
            "avail_loss":    day_metrics["avail_loss"],
            "cav_loss":      day_metrics["cav_loss"],
            "speed_loss":    day_metrics["speed_loss"],
            "quality_loss":  day_metrics["quality_loss"],
            "total_actual":  day_metrics["total_actual"],
            "total_target":  day_metrics["total_target"],
        })

    if not all_records:
        log.error("No valid data could be read from any file.")
        sys.exit(1)

    # ── Aggregate across all dates ─────────────────────────────────────────
    log.info(f"Computing aggregate across {len(daily_data)} date(s)...")
    aggregate = compute_metrics(all_records, all_not_run, all_slots)

    # ── Machine-level aggregation ──────────────────────────────────────────
    machine_nos = sorted(set(r["machine_no"] for r in all_records))
    machines = []
    for m in machine_nos:
        m_recs  = [r for r in all_records if r["machine_no"] == m]
        m_nr    = [r for r in all_not_run  if r["machine_no"] == m]
        m_slots = [r for r in all_slots    if r["machine_no"] == m]
        mc = compute_metrics(m_recs, m_nr, m_slots)

        products = [r["product_name"] for r in m_recs if r["product_name"]]
        primary_product = max(set(products), key=products.count) if products else ""
        prod_types = [r["product_type"] for r in m_recs if r["product_type"]]
        primary_type = max(set(prod_types), key=prod_types.count) if prod_types else ""

        machines.append({
            "machine_no":       m,
            "primary_product":  primary_product,
            "product_type":     primary_type,
            "overall_oee":      mc["overall_oee"],
            "oee_running":      mc["oee_running"],
            "shift_run_pct":    mc["shift_run_pct"],
            "avail_loss":       mc["avail_loss"],
            "cav_loss":         mc["cav_loss"],
            "speed_loss":       mc["speed_loss"],
            "quality_loss":     mc["quality_loss"],
            "total_actual":     mc["total_actual"],
            "total_target":     mc["total_target"],
            "total_rej":        mc["total_rej"],
            "ran_slots":        mc["ran_slots"],
            "total_slots":      mc["total_slots"],
            "dt_breakdown_hrs": mc["dt_breakdown_hrs"],
        })

    # ── Build and write JSON ───────────────────────────────────────────────
    date_range = {
        "from": daily_data[0]["date"] if daily_data else "",
        "to":   daily_data[-1]["date"] if daily_data else "",
        "days": len(daily_data),
    }

    output = {
        "_meta": {
            "generated":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_files": len(daily_data),
            "failed_files": failed_files,
            "date_range":   date_range,
            "formula_notes": {
                "overall_oee":  "OEE_running x shift_run_pct  (Sheet1 D10)",
                "oee_running":  "sum(actual) / sum(target)    (Sheet1 R6)",
                "shift_run_pct":"ran_slots / total_slots       (Sheet1 D4)",
                "avail_loss":   "(shift_hrs - run_hrs) / shift_hrs  (Sheet1 AI6)",
                "cav_loss":     "(total_cav - run_cav) / total_cav  (Sheet1 AK6)",
                "quality_loss": "rej / (rej + good)                 (Sheet1 X6)",
                "speed_loss":   "residual: 1 - oee_running/(avail*cav*quality)  (Sheet1 D7)",
                "downtime_pct": "each category hrs / total shift hrs  (Sheet2)",
                "not_run_pct":  "count / total machine-slots          (Sheet2)",
            },
        },
        "date_range": date_range,
        "aggregate":  aggregate,
        "daily":      daily_data,
        "machines":   machines,
        "records":    all_records,
        "not_run":    all_not_run,
        "all_slots":  all_slots,
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, default=str)

    log.info("Done. Written to: " + str(out_path.resolve()))
    log.info("  Dates:   " + date_range['from'] + "  to  " + date_range['to'])
    log.info("  Records: " + str(len(all_records)) + " rows across " + str(len(machines)) + " machines")

    if failed_files:
        log.warning("  Skipped " + str(len(failed_files)) + " file(s): " + str(failed_files))

    d = aggregate
    sr = d["shift_run_pct"]
    av = 1 - d["avail_loss"]
    cv = 1 - d["cav_loss"]
    sp = 1 - d["speed_loss"]
    ql = 1 - d["quality_loss"]
    print("\n" + "="*55)
    print("  WATERFALL  (" + date_range['from'] + "  to  " + date_range['to'] + ")")
    print("="*55)
    print(f"  Theoretical max          100.00%")
    print(f"  x Shift run %          {sr*100:8.2f}%   loss {(1-sr)*100:.2f}%  ({d['not_ran_slots']}/{d['total_slots']} slots idle)")
    print(f"  x Availability         {sr*av*100:8.2f}%   loss {d['avail_loss']*100:.2f}%  (downtime)")
    print(f"  x Cavity perf          {sr*av*cv*100:8.2f}%   loss {d['cav_loss']*100:.2f}%  (fewer cavities)")
    print(f"  x Speed (residual)     {sr*av*cv*sp*100:8.2f}%   loss {d['speed_loss']*100:.2f}%")
    print(f"  x Quality              {sr*av*cv*sp*ql*100:8.2f}%   loss {d['quality_loss']*100:.2f}%  (rejections)")
    print(f"  = Overall OEE          {d['overall_oee']*100:8.2f}%")
    print("="*55)

if __name__ == '__main__':
    main()
